# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime, sys

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

PERCENTILES = ['5th','10th','15th','20th','25th','30th','35th','40th','45th','50th',
               '55th','60th','65th','70th','75th','80th','85th','90th','95th']

NORM_BEND = {
    ('男',10):[11,15,16,18,19,20,21,23,24,25,25,27,27,29,30,31,32,34,36],
    ('男',11):[10,13,15,17,18,19,20,22,23,24,25,26,27,28,29,30,32,34,36],
    ('男',12):[ 9,12,15,16,17,19,20,21,22,23,24,25,26,27,29,30,31,33,36],
    ('女',10):[16,19,21,23,24,25,27,28,29,30,30,31,32,33,35,36,37,40,42],
    ('女',11):[14,18,20,22,24,25,26,27,28,29,30,31,32,33,34,35,37,39,43],
    ('女',12):[13,16,19,21,23,24,25,27,28,29,30,31,32,33,35,36,38,40,44],
}
NORM_SITUP = {
    ('男',10):[11,15,17,19,21,22,24,25,26,27,29,30,31,32,34,36,37,40,43],
    ('男',11):[14,17,20,22,23,25,26,27,29,30,31,32,34,35,36,38,40,42,46],
    ('男',12):[17,21,23,25,27,28,30,31,32,33,35,36,37,39,40,42,44,46,50],
    ('女',10):[10,14,16,18,19,21,22,23,24,25,26,28,29,30,31,33,35,37,40],
    ('女',11):[13,17,19,21,22,23,25,26,27,28,29,30,31,32,34,35,37,39,42],
    ('女',12):[16,19,21,23,24,26,27,28,29,30,31,32,33,35,36,37,39,41,44],
}
NORM_JUMP = {
    ('男',10):[100,105,110,115,119,121,125,127,130,132,135,138,141,145,148,152,156,162,170],
    ('男',11):[106,113,118,125,128,131,135,138,141,144,146,150,153,156,160,164,169,174,182],
    ('男',12):[112,122,129,133,136,141,145,148,152,155,158,161,165,169,172,176,181,187,198],
    ('女',10):[ 95,100,104,107,110,112,115,118,120,123,125,128,130,133,136,140,145,152,160],
    ('女',11):[100,105,110,114,117,120,123,125,128,131,134,137,140,142,146,150,155,160,170],
    ('女',12):[102,108,113,116,120,123,126,129,131,135,139,142,144,147,150,156,162,167,176],
}
NORM_RUN = {  # lower is better
    ('男',10):[410,385,372,360,347,336,328,318,310,300,291,283,276,267,260,251,241,231,220],
    ('男',11):[393,362,346,332,322,314,304,294,287,280,270,263,256,249,242,235,228,219,210],
    ('男',12):[361,334,319,308,297,287,279,271,264,257,250,243,237,230,223,218,212,205,194],
    ('女',10):[413,389,374,363,353,346,338,328,322,314,307,303,295,287,281,273,265,251,240],
    ('女',11):[387,363,348,338,329,322,316,309,303,296,289,283,276,269,262,255,249,236,226],
    ('女',12):[368,346,332,322,315,309,303,296,289,284,280,273,267,261,255,249,243,234,223],
}

PR_LABELS = ['<5th','5th','10th','15th','20th','25th','30th','35th','40th','45th',
             '50th','55th','60th','65th','70th','75th','80th','85th','90th','≥95th']

def get_pr(score, norm_vals, higher_better=True):
    if pd.isna(score): return ''
    try:
        score = float(score)
    except:
        return ''
    if higher_better:
        idx = 0
        for i, v in enumerate(norm_vals):
            if score >= v:
                idx = i + 1
        return PR_LABELS[idx]
    else:
        idx = 0
        for i, v in enumerate(norm_vals):
            if score <= v:
                idx = i + 1
        return PR_LABELS[idx]

def run_to_seconds(v):
    """Convert m.ss format (e.g. 3.17 = 3min17sec) to seconds."""
    if v is None or (isinstance(v, float) and v != v):
        return None
    try:
        f = float(v)
        mins = int(f)
        secs = round((f - mins) * 100)
        return mins * 60 + secs
    except:
        return None

def calc_age(birth_roc, test_date=datetime.date(2026, 3, 1)):
    try:
        s = str(int(birth_roc))
        if len(s) == 7:
            y = int(s[:3]) + 1911
            m_val = int(s[3:5])
            d_val = int(s[5:7])
            bd = datetime.date(y, m_val, d_val)
            return (test_date - bd).days // 365
    except:
        pass
    return None

def pr_color(label):
    """Return hex fill color based on PR label."""
    num_map = {'<5th':2,'5th':5,'10th':10,'15th':15,'20th':20,'25th':25,
               '30th':30,'35th':35,'40th':40,'45th':45,'50th':50,
               '55th':55,'60th':60,'65th':65,'70th':70,'75th':75,
               '80th':80,'85th':85,'90th':90,'≥95th':95}
    n = num_map.get(label, 0)
    if n == 0: return None
    if n < 25: return 'FFCCCC'   # 紅
    if n < 50: return 'FFEECC'   # 橙黃
    if n < 75: return 'FFFFCC'   # 黃
    return 'CCFFCC'              # 綠

GRADE_AGE = {'4': 10, '5': 11, '6': 12}

# === 讀取資料 ===
XLSM = r'G:\我的雲端硬碟\🎾體育組\114學年度體育組\114學年度體適能檢測\體適能檢測各班資料\114學年度體適能檢測資料各班上傳資料檔 (1).xlsm'
xf = pd.ExcelFile(XLSM, engine='openpyxl')

all_rows = []
for sheet in xf.sheet_names:
    if sheet[0] not in '456':
        continue
    df = pd.read_excel(XLSM, sheet_name=sheet, header=0, engine='openpyxl')
    col_names = [str(c).strip() for c in df.columns.tolist()]
    has_birth = '出生年月日' in col_names

    for _, row in df.iterrows():
        vals = row.tolist()

        def col(name):
            try: return vals[col_names.index(name)]
            except ValueError: return None

        try:
            int(float(vals[0]))
        except:
            continue

        birth = col('出生年月日') if has_birth else None
        age = calc_age(birth) if birth is not None else GRADE_AGE.get(sheet[0])
        g = str(col('性別') or '').strip()
        key = (g, age) if age else None

        def pr(score, norm, hb=True):
            if key and key in norm:
                return get_pr(score, norm[key], hb)
            return ''

        bend  = col('坐姿體前彎')
        jump  = col('立定跳遠')
        situp = col('仰臥起坐')
        run_raw = col('800公尺跑走')

        all_rows.append({
            '班級': col('班級') or vals[0],
            '座號': col('座號'),
            '學號': col('學號'),
            '姓名': col('姓名'),
            '性別': g,
            '出生年月日': birth if birth is not None else '',
            '年齡': age,
            '坐姿體前彎': bend,
            '坐姿_PR': pr(bend, NORM_BEND),
            '仰臥起坐': situp,
            '仰臥起坐_PR': pr(situp, NORM_SITUP),
            '立定跳遠': jump,
            '立定跳遠_PR': pr(jump, NORM_JUMP),
            '800公尺跑走': run_raw,
            '800公尺_PR': pr(run_to_seconds(run_raw), NORM_RUN, hb=False),
        })

df_all = pd.DataFrame(all_rows)
print(f'資料筆數: {len(df_all)}')

# === 建立 Excel ===
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '4-6年級體適能常模對照'

# 標頭
HEADERS = ['班級','座號','學號','姓名','性別','出生年月日','年齡',
           '坐姿體前彎','坐姿PR','仰臥起坐','仰臥起坐PR',
           '立定跳遠','立定跳遠PR','800公尺跑走','800公尺PR']

# 欄寬設定
COL_WIDTHS = [8, 6, 10, 8, 6, 12, 6, 10, 8, 10, 10, 10, 10, 12, 10]

header_fill = PatternFill('solid', start_color='1F4E79')
header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
normal_font = Font(name='Arial', size=10)
center = Alignment(horizontal='center', vertical='center')
thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# 寫標頭
for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = thin
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 18

# 寫資料
PR_COLS = [9, 11, 13, 15]  # PR 欄的 column index（1-based）
for row_idx, row in enumerate(df_all.itertuples(index=False), 2):
    data = [row[0], row[1], row[2], row[3], row[4], row[5], row[6],
            row[7], row[8], row[9], row[10], row[11], row[12], row[13], row[14]]

    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal_font
        cell.border = thin
        cell.alignment = center
        if col_idx in PR_COLS:
            color = pr_color(str(val))
            if color:
                cell.fill = PatternFill('solid', start_color=color)

ws.freeze_panes = 'A2'

# 圖例說明
legend_row = len(df_all) + 3
ws.cell(row=legend_row, column=1, value='顏色說明：').font = Font(bold=True, name='Arial')
for i, (label, color, desc) in enumerate([
    ('FFCCCC', 'FFCCCC', '< 25th (待加強)'),
    ('FFEECC', 'FFEECC', '25th~50th (中等以下)'),
    ('FFFFCC', 'FFFFCC', '50th~75th (中等以上)'),
    ('CCFFCC', 'CCFFCC', '≥ 75th (優良)'),
]):
    c = ws.cell(row=legend_row, column=i*2+2, value=desc)
    c.fill = PatternFill('solid', start_color=label)
    c.font = Font(name='Arial', size=9)

OUT = r'G:\我的雲端硬碟\🎾體育組\114學年度體育組\114學年度體適能檢測\114學年度體適能常模對照報告.xlsx'
wb.save(OUT)
print(f'已儲存: {OUT}')
